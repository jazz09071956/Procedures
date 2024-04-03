from django.core.management.base import BaseCommand

class Command(BaseCommand):
    help = (
        "Este programa se usa para importar usuarios desde un archivo XSLX local. "
        "Espera columnas: username,codigo_sala,nombre_sala,password,email,date_joined,is_active."
    )
    SILENT, NORMAL, VERBOSE, VERY_VERBOSE = 0, 1, 2, 3

    def add_arguments(self, parser):
        # Positional arguments
        parser.add_argument("file_path", nargs=1, type=str)

    def handle(self, *args, **options):
        self.verbosity = options.get("verbosity", self.NORMAL)
        self.file_path = options["file_path"][0]
        self.prepare()
        self.main()
        self.finalize()

    def prepare(self):
        self.imported_counter = 0
        self.skipped_counter = 0

    def main(self):
        from openpyxl import load_workbook
        from ...forms import UsuarioForm

        wb = load_workbook(filename=self.file_path)
        ws = wb.worksheets[0]

        if self.verbosity >= self.NORMAL:
            self.stdout.write("=== Importando usuario ===")

        columns = ["username", "codigo_sala", "nombre_sala", "password", "email", "date_joined", "is_active"]
        rows = ws.iter_rows(min_row=2)  # skip the column captions
        for index, row in enumerate(rows, start=1):
            row_values = [cell.value for cell in row]
            row_dict = dict(zip(columns, row_values))
            form = UsuarioForm(data=row_dict)
            if form.is_valid():
                usuario = form.save()
                if self.verbosity >= self.NORMAL:
                    self.stdout.write(f" - {usuario}\n")
                self.imported_counter += 1
            else:
                if self.verbosity >= self.NORMAL:
                    self.stderr.write(f"Errores importando usuario " f"{row_dict['username']} - {row_dict['codigo_sala']}:\n")
                    self.stderr.write(f"{form.errors.as_json()}\n")
                self.skipped_counter += 1

    def finalize(self):
        if self.verbosity >= self.NORMAL:
            self.stdout.write(f"-------------------------\n")
            self.stdout.write(f"Usuario imported: {self.imported_counter}\n")
            self.stdout.write(f"Usuarios ignorados: {self.skipped_counter}\n\n")